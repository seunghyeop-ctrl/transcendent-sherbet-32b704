#!/usr/bin/env python3
"""
PromptExtractor core pipeline.

Standalone runtime rules:
- No project-folder dependency
- Settings live in OS-appropriate user application data directories
- Output data lives in the user's Documents folder by default
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import subprocess
import sys
import threading
import tempfile
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

from PIL import Image as PILImage


def get_app_support_dir() -> Path:
    explicit = os.getenv("PROMPT_EXTRACTOR_APP_DIR", "").strip()
    if explicit:
        return Path(explicit).expanduser()
    home = Path.home()
    if sys.platform == "darwin":
        return home / "Library" / "Application Support" / "PromptExtractor"
    if os.name == "nt":
        appdata = os.getenv("APPDATA", "").strip()
        if appdata:
            return Path(appdata) / "PromptExtractor"
        return home / "AppData" / "Roaming" / "PromptExtractor"
    xdg = os.getenv("XDG_CONFIG_HOME", "").strip()
    if xdg:
        return Path(xdg) / "PromptExtractor"
    return home / ".config" / "PromptExtractor"


def get_default_output_root() -> Path:
    explicit = os.getenv("PROMPT_EXTRACTOR_DEFAULT_OUTPUT_DIR", "").strip()
    if explicit:
        return Path(explicit).expanduser()
    return Path.home() / "Documents" / "프롬프트 추출기"


APP_SUPPORT_DIR = get_app_support_dir()
CONFIG_PATH = APP_SUPPORT_DIR / "config.json"
DEFAULT_CREDENTIALS_PATH = APP_SUPPORT_DIR / "credentials.json"
DEFAULT_OUTPUT_ROOT = get_default_output_root()
SECRETS_PATH = APP_SUPPORT_DIR / "secrets.json"
FRAMES_DIRNAME = "frames"
OUTPUTS_DIRNAME = "outputs"
RESULTS_META_FILENAME = "results_meta.json"
DEFAULT_HEADERS = ["카메라 워킹", "프롬프트(영문)", "내용", "링크"]
DEFAULT_WORKSHEET = "Sheet1"
KEYCHAIN_ACCOUNT = "PromptExtractor"
KEYCHAIN_SERVICE = "GEMINI_API_KEY"
DEFAULT_CONFIG = {
    "google_sheet_id": "",
    "worksheet": DEFAULT_WORKSHEET,
    "output_dir": str(DEFAULT_OUTPUT_ROOT),
    "credentials_path": str(DEFAULT_CREDENTIALS_PATH),
}
MODEL_CANDIDATES = ["gemini-2.5-flash"]
DOWNLOAD_USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)
LAST_DOWNLOAD_ERROR = ""
GEMINI_PRICING_USD_PER_1M = {
    "gemini-2.5-flash": {"input": 0.30, "output": 2.50},
    "gemini-2.0-flash": {"input": 0.10, "output": 0.40},
    "gemini-2.0-flash-lite": {"input": 0.075, "output": 0.30},
}
USD_TO_KRW = 1360.0
THUMBNAIL_NAME = "keyframe.jpg"
CAMERA_MOVEMENTS = [
    "Static Shot",
    "Pan Left",
    "Pan Right",
    "Tilt Up",
    "Tilt Down",
    "Dolly In",
    "Dolly Out",
    "Truck Left",
    "Truck Right",
    "Pedestal Up",
    "Pedestal Down",
    "Zoom In",
    "Zoom Out",
    "Push In",
    "Pull Out",
    "Orbit / Arc Shot",
    "Crane Up",
    "Crane Down",
    "Handheld",
    "Steadicam",
    "Whip Pan",
    "Dutch Angle",
    "Ken Burns Effect",
    "Rack Focus",
    "Tracking Shot",
    "Follow Shot",
    "360 Rotation",
    "Drone Shot",
    "POV Shot",
    "Canted Angle",
]

ENV_CONFIG_KEYS = {
    "PROMPT_EXTRACTOR_GOOGLE_SHEET_ID": "google_sheet_id",
    "PROMPT_EXTRACTOR_WORKSHEET": "worksheet",
    "PROMPT_EXTRACTOR_OUTPUT_DIR": "output_dir",
    "PROMPT_EXTRACTOR_CREDENTIALS_PATH": "credentials_path",
}

COMMAND_CANDIDATES = {
    "yt-dlp": [
        "/Library/Frameworks/Python.framework/Versions/3.14/bin/yt-dlp",
        "/opt/homebrew/bin/yt-dlp",
        "/usr/local/bin/yt-dlp",
        "/usr/bin/yt-dlp",
    ],
    "ffmpeg": [
        "/opt/homebrew/bin/ffmpeg",
        "/usr/local/bin/ffmpeg",
        "/usr/bin/ffmpeg",
    ],
    "ffprobe": [
        "/opt/homebrew/bin/ffprobe",
        "/usr/local/bin/ffprobe",
        "/usr/bin/ffprobe",
    ],
}


def bootstrap_environment() -> None:
    APP_SUPPORT_DIR.mkdir(parents=True, exist_ok=True)
    DEFAULT_OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    (DEFAULT_OUTPUT_ROOT / FRAMES_DIRNAME).mkdir(parents=True, exist_ok=True)
    (DEFAULT_OUTPUT_ROOT / OUTPUTS_DIRNAME).mkdir(parents=True, exist_ok=True)
    if not CONFIG_PATH.exists():
        CONFIG_PATH.write_text(json.dumps(DEFAULT_CONFIG, ensure_ascii=False, indent=2), encoding="utf-8")

    credentials_json = os.getenv("GOOGLE_CREDENTIALS_JSON", "").strip()
    if credentials_json:
        try:
            parsed = json.loads(credentials_json)
            DEFAULT_CREDENTIALS_PATH.write_text(
                json.dumps(parsed, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        except Exception:
            pass


def environment_overrides() -> dict:
    overrides: dict[str, str] = {}
    for env_name, config_key in ENV_CONFIG_KEYS.items():
        value = os.getenv(env_name, "").strip()
        if value:
            overrides[config_key] = value
    return overrides


def load_config() -> dict:
    bootstrap_environment()
    try:
        data = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception:
        data = {}
    config = DEFAULT_CONFIG.copy()
    if isinstance(data, dict):
        for key, value in data.items():
            if value not in ("", None):
                config[key] = value
    config.update(environment_overrides())
    return config


def save_config(config: dict) -> dict:
    bootstrap_environment()
    existing: dict = {}
    if CONFIG_PATH.exists():
        try:
            raw = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
            if isinstance(raw, dict):
                existing = raw
        except Exception:
            existing = {}
    merged = DEFAULT_CONFIG.copy()
    merged.update(existing)
    merged.update({k: v for k, v in config.items() if v not in (None,)})
    merged.pop("gemini_api_key", None)
    CONFIG_PATH.write_text(json.dumps(merged, ensure_ascii=False, indent=2), encoding="utf-8")
    ensure_runtime_dirs(merged)
    return merged


def expand_path(path_value: str) -> Path:
    return Path(path_value).expanduser()


def ensure_runtime_dirs(config: dict) -> tuple[Path, Path, Path]:
    output_root = expand_path(config.get("output_dir", str(DEFAULT_OUTPUT_ROOT)))
    frames_dir = output_root / FRAMES_DIRNAME
    outputs_dir = output_root / OUTPUTS_DIRNAME
    output_root.mkdir(parents=True, exist_ok=True)
    frames_dir.mkdir(parents=True, exist_ok=True)
    outputs_dir.mkdir(parents=True, exist_ok=True)
    return output_root, frames_dir, outputs_dir


def csv_path_for(config: dict) -> Path:
    _, _, outputs_dir = ensure_runtime_dirs(config)
    return outputs_dir / "results.csv"


def xlsx_path_for(config: dict) -> Path:
    _, _, outputs_dir = ensure_runtime_dirs(config)
    return outputs_dir / "results.xlsx"


def metadata_path_for(config: dict) -> Path:
    _, _, outputs_dir = ensure_runtime_dirs(config)
    return outputs_dir / RESULTS_META_FILENAME


def credentials_path_for(config: dict) -> Path:
    return expand_path(config.get("credentials_path", str(DEFAULT_CREDENTIALS_PATH)))


def load_results_metadata(config: dict) -> dict[str, dict]:
    meta_path = metadata_path_for(config)
    if not meta_path.exists():
        return {}
    try:
        raw = json.loads(meta_path.read_text(encoding="utf-8"))
        if isinstance(raw, dict):
            return {str(key): value for key, value in raw.items() if isinstance(value, dict)}
    except Exception:
        return {}
    return {}


def save_results_metadata(config: dict, payload: dict[str, dict]) -> None:
    meta_path = metadata_path_for(config)
    meta_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def upsert_result_metadata(url: str, metadata: dict | None, config: dict) -> None:
    if not url or not isinstance(metadata, dict) or not metadata:
        return
    existing = load_results_metadata(config)
    existing[url] = metadata
    save_results_metadata(config, existing)


def delete_result_metadata(url: str, config: dict) -> None:
    if not url:
        return
    existing = load_results_metadata(config)
    if url in existing:
        existing.pop(url, None)
        save_results_metadata(config, existing)


def clear_results_metadata(config: dict) -> None:
    save_results_metadata(config, {})


def load_secret_store() -> dict[str, str]:
    if not SECRETS_PATH.exists():
        return {}
    try:
        raw = json.loads(SECRETS_PATH.read_text(encoding="utf-8"))
        if isinstance(raw, dict):
            return {str(key): str(value) for key, value in raw.items() if value not in (None, "")}
    except Exception:
        return {}
    return {}


def save_secret_store(payload: dict[str, str]) -> None:
    APP_SUPPORT_DIR.mkdir(parents=True, exist_ok=True)
    SECRETS_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def get_secret(name: str) -> str:
    return load_secret_store().get(name, "").strip()


def set_secret(name: str, value: str) -> bool:
    store = load_secret_store()
    cleaned = str(value or "").strip()
    if cleaned:
        store[name] = cleaned
    else:
        store.pop(name, None)
    save_secret_store(store)
    return True


def get_gemini_api_key() -> str:
    env_key = os.getenv("GEMINI_API_KEY", "").strip()
    if env_key:
        return env_key
    if sys.platform == "darwin":
        try:
            result = subprocess.run(
                [
                    "security",
                    "find-generic-password",
                    "-a",
                    KEYCHAIN_ACCOUNT,
                    "-s",
                    KEYCHAIN_SERVICE,
                    "-w",
                ],
                capture_output=True,
                text=True,
                check=True,
            )
            found = result.stdout.strip()
            if found:
                return found
        except (subprocess.CalledProcessError, FileNotFoundError):
            pass
    return get_secret(KEYCHAIN_SERVICE)


def set_gemini_api_key(value: str) -> bool:
    cleaned = str(value or "").strip()
    if sys.platform == "darwin":
        try:
            subprocess.run(
                [
                    "security",
                    "add-generic-password",
                    "-a",
                    KEYCHAIN_ACCOUNT,
                    "-s",
                    KEYCHAIN_SERVICE,
                    "-w",
                    cleaned,
                    "-U",
                ],
                capture_output=True,
                text=True,
                check=True,
            )
            return True
        except (subprocess.CalledProcessError, FileNotFoundError):
            pass
    return set_secret(KEYCHAIN_SERVICE, cleaned)


def has_gemini_api_key() -> bool:
    return bool(get_gemini_api_key().strip())


def get_video_id(url: str) -> str:
    parsed = urlparse(url)
    return parsed.path.rstrip("/").split("/")[-1] or "unknown_video"


def sanitize_slug(text: str, fallback: str) -> str:
    cleaned: list[str] = []
    last_sep = False
    for char in text.lower():
        if char.isalnum():
            cleaned.append(char)
            last_sep = False
        elif not last_sep:
            cleaned.append("_")
            last_sep = True
    slug = "".join(cleaned).strip("_")
    parts = [part for part in slug.split("_") if part]
    return "_".join(parts[:8]) if parts else fallback


def extract_video_id_from_name(text: str) -> str:
    parts = [part for part in str(text).replace("\\", "/").rstrip("/").split("/")[-1].split("_") if part]
    if not parts:
        return ""
    last = parts[-1]
    return last if len(last) >= 6 else ""


def was_cancelled(cancel_event: threading.Event | None) -> bool:
    return bool(cancel_event and cancel_event.is_set())


def run_command(
    cmd: list[str],
    timeout: int,
    error_message: str,
    cancel_event: threading.Event | None = None,
) -> subprocess.CompletedProcess[str] | None:
    global LAST_DOWNLOAD_ERROR
    if was_cancelled(cancel_event):
        print("⏹️  사용자 중단 요청")
        return None
    try:
        process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        start = time.time()
        while True:
            if was_cancelled(cancel_event):
                process.terminate()
                try:
                    process.wait(timeout=2)
                except subprocess.TimeoutExpired:
                    process.kill()
                print("⏹️  사용자 중단 요청")
                return None
            if process.poll() is not None:
                stdout, stderr = process.communicate()
                result = subprocess.CompletedProcess(cmd, process.returncode, stdout, stderr)
                break
            if time.time() - start > timeout:
                process.terminate()
                try:
                    process.wait(timeout=2)
                except subprocess.TimeoutExpired:
                    process.kill()
                print(f"❌ {error_message}: 시간 초과")
                if error_message.startswith("다운로드 실패"):
                    LAST_DOWNLOAD_ERROR = "시간 초과"
                return None
            time.sleep(0.2)
    except FileNotFoundError:
        print(f"❌ 명령어를 찾을 수 없습니다: {cmd[0]}")
        if error_message.startswith("다운로드 실패"):
            LAST_DOWNLOAD_ERROR = f"명령어를 찾을 수 없습니다: {cmd[0]}"
        return None

    if result.returncode != 0:
        detail = (result.stderr or result.stdout or "원인 미상").strip()
        print(f"❌ {error_message}: {detail}")
        if error_message.startswith("다운로드 실패"):
            LAST_DOWNLOAD_ERROR = detail
        return None
    return result


def resolve_command(name: str) -> str:
    direct = shutil.which(name)
    if direct:
        return direct
    for candidate in COMMAND_CANDIDATES.get(name, []):
        if Path(candidate).exists():
            return candidate
    return name


def sample_frames(frame_paths: list[Path], limit: int) -> list[Path]:
    if len(frame_paths) <= limit:
        return frame_paths
    indexes = sorted({round(i * (len(frame_paths) - 1) / (limit - 1)) for i in range(limit)})
    return [frame_paths[index] for index in indexes]


def normalize_download_url(url: str) -> str:
    parsed = urlparse(url.strip())
    if "instagram.com" in parsed.netloc.lower():
        return parsed._replace(query="", fragment="").geturl()
    return url.strip()


def write_download_cookies(output_dir: Path) -> Path | None:
    cookies_path = os.getenv("YTDLP_COOKIES_FILE", "").strip()
    if cookies_path and Path(cookies_path).expanduser().exists():
        return Path(cookies_path).expanduser()
    cookies_text = os.getenv("YTDLP_COOKIES", "").strip() or os.getenv("INSTAGRAM_COOKIES", "").strip()
    if not cookies_text:
        return None
    target = output_dir / "cookies.txt"
    target.write_text(cookies_text, encoding="utf-8")
    return target


def compact_download_error(detail: str) -> str:
    cleaned = re.sub(r"\s+", " ", detail or "").strip()
    if not cleaned:
        return "원인 미상"
    lowered = cleaned.lower()
    if "login" in lowered or "cookies" in lowered or "not logged in" in lowered:
        return "인스타그램이 로그인/쿠키 인증을 요구했습니다. Render 환경변수에 YTDLP_COOKIES를 설정해야 할 수 있습니다."
    if "private" in lowered or "unavailable" in lowered:
        return "영상이 비공개이거나 현재 다운로드할 수 없는 상태입니다."
    if "429" in lowered or "rate" in lowered or "blocked" in lowered:
        return "인스타그램이 Render 서버의 다운로드 요청을 제한했습니다. 잠시 뒤 재시도하거나 쿠키 설정이 필요합니다."
    return cleaned[:260]


def download_video(url: str, output_dir: Path, cancel_event: threading.Event | None = None) -> Path | None:
    global LAST_DOWNLOAD_ERROR
    LAST_DOWNLOAD_ERROR = ""
    output_dir.mkdir(parents=True, exist_ok=True)
    source_url = normalize_download_url(url)
    cookies_file = write_download_cookies(output_dir)
    base_cmd = [
        resolve_command("yt-dlp"),
        "--no-playlist",
        "--retries",
        "3",
        "--fragment-retries",
        "3",
        "--force-ipv4",
        "--user-agent",
        DOWNLOAD_USER_AGENT,
        "--referer",
        "https://www.instagram.com/",
        "--add-header",
        "Accept-Language: ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        "--output",
        str(output_dir / "video.%(ext)s"),
        "--no-warnings",
        "--print",
        "after_move:filepath",
    ]
    if cookies_file:
        base_cmd.extend(["--cookies", str(cookies_file)])
    attempts = [
        base_cmd + ["--format", "mp4/bestvideo[ext=mp4]+bestaudio[ext=m4a]/best", source_url],
        base_cmd + ["--format", "best[ext=mp4]/best", source_url],
    ]
    print("⬇️  다운로드 중...")
    last_detail = ""
    for index, cmd in enumerate(attempts, start=1):
        result = run_command(cmd, timeout=240, error_message=f"다운로드 실패 ({index}/{len(attempts)})", cancel_event=cancel_event)
        if result:
            filepath = result.stdout.strip().splitlines()[-1] if result.stdout.strip() else ""
            if filepath and Path(filepath).exists():
                print("✅ 다운로드 완료")
                return Path(filepath)
            videos = sorted(output_dir.glob("video.*"))
            if videos:
                print("✅ 다운로드 완료")
                return videos[0]
        last_detail = LAST_DOWNLOAD_ERROR or last_detail
    videos = sorted(output_dir.glob("video.*"))
    if videos:
        print("✅ 다운로드 완료")
        return videos[0]
    LAST_DOWNLOAD_ERROR = compact_download_error(last_detail)
    return None


def get_duration(video_path: Path, cancel_event: threading.Event | None = None) -> float:
    cmd = [
        resolve_command("ffprobe"),
        "-v",
        "error",
        "-show_entries",
        "format=duration",
        "-of",
        "default=noprint_wrappers=1:nokey=1",
        str(video_path),
    ]
    result = run_command(cmd, timeout=10, error_message="영상 길이 확인 실패", cancel_event=cancel_event)
    if not result:
        return 30.0
    try:
        return float(result.stdout.strip())
    except ValueError:
        return 30.0


def extract_frames(video_path: Path, output_dir: Path, cancel_event: threading.Event | None = None) -> list[Path]:
    duration = get_duration(video_path, cancel_event=cancel_event)
    if was_cancelled(cancel_event):
        return []
    output_dir.mkdir(parents=True, exist_ok=True)
    ffmpeg_cmd = resolve_command("ffmpeg")

    extra_tail_frames = 0
    tail_seconds = 0.0
    if duration <= 20:
        fps = "2"
        max_frames = 30
        extra_tail_frames = 8
        tail_seconds = min(4.0, duration)
    elif duration <= 60:
        fps = "1"
        max_frames = 48
        extra_tail_frames = 10
        tail_seconds = min(6.0, duration)
    else:
        fps = "2/3"
        max_frames = 48
        extra_tail_frames = 8
        tail_seconds = min(8.0, duration)

    cmd = [
        ffmpeg_cmd,
        "-i",
        str(video_path),
        "-vf",
        f"fps={fps}",
        "-q:v",
        "2",
        "-frames:v",
        str(max_frames),
        str(output_dir / "frame_%03d.jpg"),
        "-y",
        "-loglevel",
        "error",
    ]
    print(f"🎞️  프레임 추출 중 (영상 {duration:.0f}초, 기본 {max_frames}장 + 후반 보강 {extra_tail_frames}장)...")
    result = run_command(cmd, timeout=180, error_message="프레임 추출 실패", cancel_event=cancel_event)
    if not result:
        return []

    frames = sorted(output_dir.glob("frame_*.jpg"))
    if was_cancelled(cancel_event):
        return []

    if extra_tail_frames > 0 and tail_seconds > 0 and duration > tail_seconds:
        tail_start = max(duration - tail_seconds, 0.0)
        tail_cmd = [
            ffmpeg_cmd,
            "-ss",
            f"{tail_start:.3f}",
            "-i",
            str(video_path),
            "-vf",
            f"fps={extra_tail_frames / max(tail_seconds, 1):.6f}",
            "-q:v",
            "2",
            "-frames:v",
            str(extra_tail_frames),
            "-start_number",
            str(len(frames) + 1),
            str(output_dir / "frame_%03d.jpg"),
            "-y",
            "-loglevel",
            "error",
        ]
        tail_result = run_command(tail_cmd, timeout=180, error_message="후반 프레임 보강 실패", cancel_event=cancel_event)
        if tail_result:
            frames = sorted(output_dir.glob("frame_*.jpg"))

    print(f"✅ 프레임 {len(frames)}장 추출 완료")
    return frames


def build_ocr_crop_paths(frame_path: Path, detail_results: list | None = None) -> list[Path]:
    crop_paths: list[Path] = []
    try:
        with PILImage.open(frame_path) as image:
            image = image.convert("RGB")
            width, height = image.size
            crop_boxes: list[tuple[int, int, int, int]] = []

            if detail_results:
                candidates: list[tuple[int, int, int, int, int]] = []
                for item in detail_results:
                    try:
                        points = item[0]
                        line = " ".join(str(item[1]).split()).strip()
                    except Exception:
                        continue
                    if not line:
                        continue
                    score = score_prompt_like_line(line)
                    if score <= 0:
                        continue
                    xs = [int(pt[0]) for pt in points]
                    ys = [int(pt[1]) for pt in points]
                    left, top, right, bottom = min(xs), min(ys), max(xs), max(ys)
                    if right - left < 30 or bottom - top < 10:
                        continue
                    candidates.append((left, top, right, bottom, score))

                candidates.sort(key=lambda item: item[4], reverse=True)
                merged_regions: list[list[int]] = []
                for left, top, right, bottom, score in candidates:
                    expanded = [
                        max(0, left - int(width * 0.04)),
                        max(0, top - int(height * 0.04)),
                        min(width, right + int(width * 0.04)),
                        min(height, bottom + int(height * 0.04)),
                        score,
                    ]
                    attached = False
                    for region in merged_regions:
                        overlap_x = not (expanded[2] < region[0] or expanded[0] > region[2])
                        overlap_y = not (expanded[3] < region[1] or expanded[1] > region[3])
                        close_x = abs(expanded[0] - region[2]) < width * 0.08 or abs(expanded[2] - region[0]) < width * 0.08
                        close_y = abs(expanded[1] - region[3]) < height * 0.08 or abs(expanded[3] - region[1]) < height * 0.08
                        if (overlap_x and overlap_y) or (close_x and overlap_y) or (close_y and overlap_x):
                            region[0] = min(region[0], expanded[0])
                            region[1] = min(region[1], expanded[1])
                            region[2] = max(region[2], expanded[2])
                            region[3] = max(region[3], expanded[3])
                            region[4] += score
                            attached = True
                            break
                    if not attached:
                        merged_regions.append(expanded)

                merged_regions.sort(key=lambda item: (item[4], (item[2]-item[0]) * (item[3]-item[1])), reverse=True)
                for left, top, right, bottom, _score in merged_regions[:3]:
                    crop_boxes.append((left, top, right, bottom))

            if not crop_boxes:
                crop_boxes = [
                    (0, 0, width, height),
                    (int(width * 0.18), int(height * 0.18), int(width * 0.82), int(height * 0.82)),
                    (0, int(height * 0.55), width, height),
                ]
            else:
                crop_boxes.insert(0, (0, 0, width, height))

            unique_boxes: list[tuple[int, int, int, int]] = []
            seen_boxes: set[tuple[int, int, int, int]] = set()
            for box in crop_boxes:
                normalized = tuple(int(v) for v in box)
                if normalized in seen_boxes:
                    continue
                seen_boxes.add(normalized)
                unique_boxes.append(normalized)

            for idx, box in enumerate(unique_boxes, start=1):
                cropped = image.crop(box)
                if cropped.width < 50 or cropped.height < 50:
                    continue
                scale = 1 if idx == 1 else 2
                enlarged = cropped.resize((cropped.width * scale, cropped.height * scale))
                handle = tempfile.NamedTemporaryFile(prefix=f"ocr_crop_{frame_path.stem}_{idx}_", suffix=".png", delete=False)
                handle.close()
                out_path = Path(handle.name)
                enlarged.save(out_path, format="PNG")
                crop_paths.append(out_path)
    except Exception:
        return []
    return crop_paths


def cleanup_temp_paths(paths: list[Path]) -> None:
    for path in paths:
        try:
            path.unlink(missing_ok=True)
        except Exception:
            pass


def run_ocr_all_frames(frame_paths: list[Path], cancel_event: threading.Event | None = None) -> list[dict]:
    try:
        import easyocr
    except ImportError:
        print("⚠️  easyocr가 설치되어 있지 않습니다.")
        return []

    print(f"🔍 OCR 수행 중 ({len(frame_paths)}장 전체 분석)...")
    reader = easyocr.Reader(["en", "ko"], gpu=False)
    results: list[dict] = []
    for index, frame_path in enumerate(frame_paths, start=1):
        if was_cancelled(cancel_event):
            print("⏹️  OCR 중 사용자 중단")
            break

        collected: list[str] = []
        seen: set[str] = set()
        crop_paths: list[Path] = []
        try:
            try:
                full_detail = reader.readtext(str(frame_path), detail=1)
            except Exception:
                full_detail = []

            for item in full_detail:
                try:
                    line = item[1]
                except Exception:
                    continue
                clean = " ".join(str(line).split()).strip()
                if not clean or clean in seen:
                    continue
                seen.add(clean)
                collected.append(clean)

            crop_paths = build_ocr_crop_paths(frame_path, full_detail)
            for target in crop_paths:
                try:
                    lines = reader.readtext(str(target), detail=0)
                except Exception:
                    lines = []
                for line in lines:
                    clean = " ".join(str(line).split()).strip()
                    if not clean or clean in seen:
                        continue
                    seen.add(clean)
                    collected.append(clean)
        except Exception as exc:
            print(f"⚠️  OCR 실패: {frame_path.name} ({exc})")
        finally:
            cleanup_temp_paths(crop_paths)

        results.append({"frame": frame_path.name, "lines": collected})
        if index % 5 == 0 or index == len(frame_paths):
            print(f"   {index}/{len(frame_paths)}장 처리")
    return results


def score_prompt_like_line(raw_line: str) -> int:
    line = " ".join(str(raw_line).split()).strip()
    if not line:
        return -10
    lower = line.lower()

    blocked_fragments = (
        "follow",
        "like",
        "comment",
        "share",
        "subscribe",
        "instagram",
        "reel",
        "original audio",
        "사용방법",
        "궁금하신분들",
        "이거 보셨나요",
        "4일만에 만든",
        "ai 영화",
        "higgsfield",
        "character location",
    )
    if any(fragment in lower for fragment in blocked_fragments):
        return -10

    score = 0
    if any(keyword in lower for keyword in (
        "shot", "camera", "lens", "lighting", "depth of field", "wide", "close-up",
        "tracking", "steadicam", "handheld", "low-angle", "dynamic", "scene",
        "time-freeze", "time freeze", "cinematic", "ultrarealistic", "arri", "prompt",
    )):
        score += 4
    if any(keyword in lower for keyword in (
        "character", "interior", "exterior", "window", "monster", "restaurant", "beer",
        "arena", "creature", "hero", "giant", "glass", "shockwave", "walking", "frozen",
    )):
        score += 2
    if ':' in line or ';' in line or '[' in line or ']' in line:
        score += 1
    if len(line.split()) >= 6:
        score += 2
    if sum('가' <= ch <= '힣' for ch in line) > max(3, len(line) // 4):
        score -= 3
    if line.startswith('@'):
        score -= 2
    return score


def select_best_ocr_window(ocr_results: list[dict], window_size: int = 3) -> list[dict]:
    if not ocr_results:
        return []

    frame_scores: list[int] = []
    for item in ocr_results:
        lines = item.get('lines', [])
        score = sum(max(score_prompt_like_line(line), 0) for line in lines)
        frame_scores.append(score)

    best_total = -1
    best_slice = (0, min(len(ocr_results), window_size))
    for start in range(len(ocr_results)):
        end = min(len(ocr_results), start + window_size)
        total = sum(frame_scores[start:end])
        if total > best_total:
            best_total = total
            best_slice = (start, end)

    start, end = best_slice
    selected = ocr_results[start:end]
    if best_total <= 0:
        return ocr_results
    return selected


def aggregate_ocr_text(ocr_results: list[dict]) -> str:
    seen: set[str] = set()
    merged: list[str] = []
    fallback_all: list[str] = []

    for item in ocr_results:
        for raw_line in item.get("lines", []):
            clean = " ".join(str(raw_line).split()).strip()
            if not clean or clean in seen:
                continue
            seen.add(clean)
            fallback_all.append(clean)
            if score_prompt_like_line(clean) > 0:
                merged.append(clean)

    if merged:
        return "\n".join(merged).strip()
    return "\n".join(fallback_all).strip()


def fallback_prompt_lines_from_ocr(ocr_results: list[dict], limit: int = 40) -> list[str]:
    seen: set[str] = set()
    selected: list[str] = []
    fallback_all: list[str] = []
    for item in ocr_results:
        for raw_line in item.get("lines", []):
            line = " ".join(str(raw_line).split()).strip()
            if not line or line in seen:
                continue
            if len(line) < 4:
                continue
            seen.add(line)
            fallback_all.append(line)
            if score_prompt_like_line(line) > 0:
                selected.append(line)
                if len(selected) >= limit:
                    return selected
    return selected if selected else fallback_all[:limit]


def detect_camera_movement(prompt: str) -> str:
    text = (prompt or "").lower()
    if not text:
        return ""

    keyword_map = [
        ([("drone aerial shot",), ("drone shot", "aerial")], "Drone aerial shot"),
        ([("extreme close-up",)], "Extreme close-up"),
        ([("low angle", "static"), ("low-angle", "static")], "Low angle static"),
        ([("handheld", "vibrat")], "Handheld vibration"),
        ([("close-up", "tracking"), ("close up", "tracking")], "Close-up tracking"),
        ([("steadicam", "orbit"), ("steadicam", "orbital")], "Steadicam orbital"),
        ([("wide shot", "handheld"), ("handheld", "wide shot")], "Handheld wide shot"),
        ([("low-angle", "tracking"), ("low angle", "tracking")], "Low-angle tracking"),
        ([("imax", "orbit"), ("imax", "orbital")], "IMAX orbital"),
        ([("dynamic", "low-angle"), ("dynamic", "low angle")], "Dynamic low-angle"),
        ([("pulls back",), ("pull back",), ("camera pulls back",)], "Camera pulls back"),
        ([("camera slowly pans",), ("camera pans",), ("pan shot",)], "Pan shot"),
        ([("steadicam", "tracking"), ("front tracking shot",), ("tracking shot", "steadicam"), ("steadicam", "flying"), ("continuous long take", "tracking shot"), ("one continuous long take", "tracking shot")], "Steadicam tracking"),
        ([("tracking shot",), ("tracking",)], "Tracking shot"),
        ([("flying backward",), ("camera continues flying backward",), ("walks toward the lens",), ("toward the lens",)], "Tracking shot"),
        ([("flying backward",), ("camera flying backward",), ("camera from medium shot flying backward",)], "Dolly Out"),
        ([("one continuous long take",), ("continuous long take",), ("no cuts whatsoever",)], "Long take"),
        ([("handheld",)], "Handheld shot"),
        ([("wide shot",)], "Wide shot"),
        ([("close-up",), ("close up",)], "Close-up"),
        ([("dolly in",), ("dolly push",)], "Dolly In"),
        ([("dolly out",), ("dolly back",)], "Dolly Out"),
        ([("zoom in",)], "Zoom In"),
        ([("zoom out",)], "Zoom Out"),
        ([("tilt up",)], "Tilt Up"),
        ([("tilt down",)], "Tilt Down"),
    ]
    matched: list[str] = []
    for groups, label in keyword_map:
        for needles in groups:
            if all(needle in text for needle in needles):
                if label not in matched:
                    matched.append(label)
                break
    return ' / '.join(matched)


def build_korean_summary(raw_row: dict) -> str:
    source = (raw_row.get("KO") or raw_row.get("PROMPT") or "").strip()
    if not source:
        return ""

    text = source.replace("\n", " ").replace("—", " ")
    text = " ".join(text.split())
    lower = text.lower()

    preset_rules = [
        (("racing track", "time has stopped"), "밤의 레이싱 트랙 및 시간 정지 연출"),
        (("racing track", "f1"), "밤의 레이싱 트랙 및 시간 정지 연출"),
        (("coconut", "cat"), "코코넛 낙하 및 시간 정지 반전 연출"),
        (("card", "masked men"), "도박장 카드 투척 및 격투 시퀀스 (세부설정 포함)"),
        (("leviathan", "bridge"), "부서진 다리 위 물 괴수와의 전투"),
        (("mentor", "takeoff"), "멘토의 격려와 비행기 이륙 시퀀스"),
        (("raindrop", "time freeze"), "시간 정지 속 빗방울 터치 상세 연출"),
        (("glass bottle", "bar/restaurant"), "바 앞 유리병 투척 및 반사신경 연출"),
        (("restaurant", "beer"), "레스토랑에서 맥주를 가로채는 시간 정지 연출"),
        (("restaurant", "waiter", "tray"), "레스토랑에서 맥주를 가로채는 시간 정지 연출"),
        (("boss monster",), "우주 전사와 보스 몬스터의 대치"),
        (("demon god", "awakens"), "주인공의 고통과 마신으로의 각성"),
        (("stadium", "alien"), "경기장 질주 및 외계인 관찰"),
        (("hotel", "shockwave"), "호텔 앞 시간 정지 및 충격파 연출"),
        (("pigeons", "woman"), "시간 정지 속 여인과 비둘기 연출"),
        (("cyclopean", "window"), "거실 창밖 외눈 괴물과의 대치"),
    ]
    for needles, summary in preset_rules:
        if all(needle in lower for needle in needles):
            return summary

    tokens = set()
    noun_map = [
        ("time freeze", "시간 정지"),
        ("time-freeze", "시간 정지"),
        ("racing track", "레이싱 트랙"),
        ("night", "밤"),
        ("coconut", "코코넛"),
        ("cat", "고양이"),
        ("card", "카드"),
        ("masked men", "도박장"),
        ("bridge", "다리 위"),
        ("leviathan", "물 괴수"),
        ("mentor", "멘토"),
        ("takeoff", "이륙"),
        ("raindrop", "빗방울"),
        ("hotel", "호텔 앞"),
        ("shockwave", "충격파"),
        ("woman", "여인"),
        ("pigeons", "비둘기"),
        ("glass bottle", "유리병 투척"),
        ("bar", "바 앞"),
        ("boss monster", "보스 몬스터"),
        ("space warrior", "우주 전사"),
        ("demon god", "마신"),
        ("pain", "고통"),
        ("awakens", "각성"),
        ("alien", "외계인"),
        ("stadium", "경기장"),
        ("run", "질주"),
    ]
    for needle, label in noun_map:
        if needle in lower:
            tokens.add(label)

    combo_rules = [
        (("밤", "레이싱 트랙", "시간 정지"), "밤의 레이싱 트랙 및 시간 정지 연출"),
        (("코코넛", "고양이", "시간 정지"), "코코넛 낙하 및 시간 정지 반전 연출"),
        (("도박장", "카드"), "도박장 카드 투척 및 격투 시퀀스 (세부설정 포함)"),
        (("다리 위", "물 괴수"), "부서진 다리 위 물 괴수와의 전투"),
        (("멘토", "이륙"), "멘토의 격려와 비행기 이륙 시퀀스"),
        (("빗방울", "시간 정지"), "시간 정지 속 빗방울 터치 상세 연출"),
        (("바 앞", "유리병 투척"), "바 앞 유리병 투척 및 반사신경 연출"),
        (("우주 전사", "보스 몬스터"), "우주 전사와 보스 몬스터의 대치"),
        (("고통", "마신", "각성"), "주인공의 고통과 마신으로의 각성"),
        (("경기장", "질주", "외계인"), "경기장 질주 및 외계인 관찰"),
        (("호텔 앞", "시간 정지", "충격파"), "호텔 앞 시간 정지 및 충격파 연출"),
        (("시간 정지", "여인", "비둘기"), "시간 정지 속 여인과 비둘기 연출"),
    ]
    for needs, summary in combo_rules:
        if all(item in tokens for item in needs):
            return summary

    cleaned = text
    for src in ("@image1", "Image 1", "image1", "Photo1", "main character"):
        cleaned = cleaned.replace(src, "주인공")
    parts = []
    for part in cleaned.replace("?", ".").replace("!", ".").split("."):
        part = part.strip(" -:;")
        if len(part) < 6:
            continue
        low = part.lower()
        blocked = ("arri", "alexa", "lens", "35mm", "50mm", "f/", "film grain", "ultra realistic", "cinematic", "lighting", "depth of field", "close-up", "medium shot", "wide shot", "tracking shot", "steadicam", "drone shot", "camera")
        if any(fragment in low for fragment in blocked):
            continue
        parts.append(part)
    summary = parts[0] if parts else cleaned[:28]
    return summary[:28].strip(" /,-")


def build_summary_source(aggregated_ocr: str, prompt: str = "") -> str:
    ocr_lines = [" ".join(str(line).split()).strip() for line in aggregated_ocr.splitlines() if str(line).strip()]
    preferred: list[str] = []
    fallback: list[str] = []
    for line in ocr_lines:
        score = score_prompt_like_line(line)
        if score > 0:
            preferred.append(line)
        else:
            fallback.append(line)
    parts = preferred[:40] if preferred else fallback[:20]
    if prompt.strip():
        parts.append(prompt.strip())
    return "\n".join(parts).strip()


def export_row(raw_row: dict) -> dict[str, str]:
    return {
        "카메라 워킹": raw_row.get("CAMERA", ""),
        "프롬프트(영문)": raw_row.get("PROMPT", ""),
        "내용": (raw_row.get("CONTENT") or "").strip() or build_korean_summary(raw_row),
        "링크": raw_row.get("URL", ""),
    }


def estimate_gemini_cost(model_name: str, input_tokens: int, output_tokens: int) -> dict:
    pricing = GEMINI_PRICING_USD_PER_1M.get(model_name, GEMINI_PRICING_USD_PER_1M["gemini-2.0-flash"])
    usd = ((input_tokens / 1_000_000) * pricing["input"]) + ((output_tokens / 1_000_000) * pricing["output"])
    krw = usd * USD_TO_KRW
    return {
        "model": model_name,
        "input_tokens": int(input_tokens),
        "output_tokens": int(output_tokens),
        "usd_estimate": round(usd, 6),
        "krw_estimate": round(krw, 2),
    }


def extract_usage_metadata(response, model_name: str) -> dict:
    usage = getattr(response, "usage_metadata", None) or getattr(response, "usageMetadata", None)
    if not usage:
        return estimate_gemini_cost(model_name, 0, 0)
    input_tokens = getattr(usage, "prompt_token_count", None) or getattr(usage, "promptTokenCount", None) or 0
    output_tokens = getattr(usage, "candidates_token_count", None) or getattr(usage, "candidatesTokenCount", None) or 0
    return estimate_gemini_cost(model_name, input_tokens, output_tokens)


def build_video_gemini_prompt() -> str:
    return """
You are extracting on-screen AI prompt text from a short video.

Rules:
- Read the prompt text directly from the video itself.
- Preserve visible wording as faithfully as possible.
- It is okay to merge obvious line breaks or split fragments when clearly the same sentence.
- Preserve timestamps like [0:00-0:03] when visible.
- Preserve non-English dialogue lines when visible.
- Include late tail content and Sound blocks if present.
- Do not improve style or invent missing content.

Return only JSON in this schema:
{
  "prompt_lines": ["line 1", "line 2"],
  "candidate_lines": ["optional fallback line"],
  "notes": "brief note"
}
""".strip()


def wait_for_uploaded_file(client, uploaded_file, cancel_event: threading.Event | None = None, timeout_seconds: int = 180) -> tuple[object | None, str]:
    started = time.time()
    current = uploaded_file
    while time.time() - started < timeout_seconds:
        if was_cancelled(cancel_event):
            return None, "사용자 중단"
        try:
            current = client.files.get(name=uploaded_file.name)
        except Exception as exc:
            return None, f"파일 상태 조회 실패: {exc}"
        state = getattr(current, 'state', None)
        state_name = getattr(state, 'name', '') if state else ''
        if state_name == 'ACTIVE':
            return current, ''
        if state_name in {'FAILED', 'CANCELLED'}:
            return None, f"파일 처리 실패: {state_name}"
        time.sleep(2)
    return None, "파일 처리 시간 초과"


def call_gemini_on_video(video_path: Path, gemini_api_key: str, cancel_event: threading.Event | None = None) -> tuple[dict, str, dict]:
    if not gemini_api_key.strip():
        return {}, "Gemini API 키가 설정되지 않았습니다.", estimate_gemini_cost("gemini-2.5-flash", 0, 0)
    if was_cancelled(cancel_event):
        return {}, "사용자 중단", estimate_gemini_cost("gemini-2.5-flash", 0, 0)
    try:
        from google import genai
        from google.genai import types
    except ImportError:
        return {}, "google-genai 라이브러리가 설치되어 있지 않습니다.", estimate_gemini_cost("gemini-2.5-flash", 0, 0)

    client = genai.Client(api_key=gemini_api_key)
    last_error = ''
    last_usage = estimate_gemini_cost('gemini-2.5-flash', 0, 0)

    for model_name in MODEL_CANDIDATES:
        for retry in range(3):
            uploaded_file = None
            try:
                uploaded_file = client.files.upload(file=str(video_path))
                active_file, file_error = wait_for_uploaded_file(client, uploaded_file, cancel_event=cancel_event)
                if file_error:
                    last_error = f"{model_name}: {file_error}"
                    break
                response = client.models.generate_content(
                    model=model_name,
                    contents=[build_video_gemini_prompt(), active_file],
                    config=types.GenerateContentConfig(
                        response_mime_type='application/json',
                        temperature=0.1,
                    ),
                )
                last_usage = extract_usage_metadata(response, model_name)
                raw_text = (response.text or '').strip()
                if raw_text:
                    return json.loads(raw_text), '', last_usage
            except Exception as exc:
                last_error = f"{model_name}: {exc}"
                error_text = str(exc)
                if ("429" in error_text or "503" in error_text or "UNAVAILABLE" in error_text) and retry < 2:
                    time.sleep(5 * (retry + 1))
                    continue
            finally:
                if uploaded_file is not None:
                    try:
                        client.files.delete(name=uploaded_file.name)
                    except Exception:
                        pass
    return {}, last_error or '응답 없음', last_usage


def build_gemini_prompt(ocr_text: str) -> str:
    return f"""
당신은 OCR 추출 텍스트에서 프롬프트 문장만 골라내는 필터다.

핵심 규칙:
- OCR 원문에 있는 텍스트만 사용할 것
- 프롬프트로 보이는 문장만 추릴 것
- 절대로 새로운 내용을 추가하지 말 것
- 같은 문장의 조각이 명확하게 이어지는 경우에만 자연스럽게 합칠 수 있다
- 명확한 OCR 잘림만 최소한으로 복원할 수 있다
- 중복되거나 반복된 파편은 제거할 수 있다
- 의미를 바꾸는 재작성, 요약, 번역은 금지
- 순서는 최대한 유지할 것
- 불필요한 문장만 제거할 것
- 프롬프트가 없으면 빈 배열을 반환할 것
- 애매하더라도 프롬프트 문장일 가능성이 높으면 포함할 것
- 짧더라도 묘사 지시문, 스타일 키워드, 카메라/조명/구도 지시어는 유지할 것
- 화면 UI 문구, 좋아요/댓글/팔로우 유도, 계정명, 해시태그, 자막 설명은 제거할 것

OCR 전체 텍스트:
{ocr_text if ocr_text else "(없음)"}

반드시 아래 JSON 형식으로만 답하라:
{{
  "prompt_lines": [
    "원문 줄1",
    "원문 줄2"
  ],
  "candidate_lines": [
    "애매하지만 프롬프트 가능성이 높은 원문 줄"
  ]
}}
""".strip()


def call_gemini(ocr_text: str, gemini_api_key: str, cancel_event: threading.Event | None = None) -> tuple[dict, str, dict]:
    if not gemini_api_key.strip():
        return {}, "Gemini API 키가 설정되지 않았습니다.", estimate_gemini_cost("gemini-2.0-flash", 0, 0)
    if was_cancelled(cancel_event):
        return {}, "사용자 중단", estimate_gemini_cost("gemini-2.0-flash", 0, 0)
    try:
        from google import genai
        from google.genai import types
    except ImportError:
        return {}, "google-genai 라이브러리가 설치되어 있지 않습니다.", estimate_gemini_cost("gemini-2.0-flash", 0, 0)

    parts = [types.Part.from_text(text=build_gemini_prompt(ocr_text))]

    client = genai.Client(api_key=gemini_api_key)
    last_error = ""
    last_usage = estimate_gemini_cost("gemini-2.0-flash", 0, 0)
    for model_name in MODEL_CANDIDATES:
        for retry in range(3):
            if was_cancelled(cancel_event):
                return {}, "사용자 중단", last_usage
            try:
                response = client.models.generate_content(
                    model=model_name,
                    contents=parts,
                    config=types.GenerateContentConfig(
                        response_mime_type="application/json",
                        temperature=0.1,
                    ),
                )
                last_usage = extract_usage_metadata(response, model_name)
                raw_text = (response.text or "").strip()
                if raw_text:
                    return json.loads(raw_text), "", last_usage
            except Exception as exc:
                last_error = f"{model_name}: {exc}"
                error_text = str(exc)
                if ("429" in error_text or "503" in error_text or "UNAVAILABLE" in error_text) and retry < 2:
                    time.sleep(5 * (retry + 1))
                    continue
    return {}, last_error or "응답 없음", last_usage


def clean_prompt_line(line: str) -> str:
    line = re.sub(r"\s+", " ", line).strip()
    line = re.sub(r"\b([A-Za-z0-9@#]+)(\s+\1\b)+", r"\1", line, flags=re.IGNORECASE)
    line = re.sub(r"\b(\d+mm)(\s+\1\b)+", r"\1", line, flags=re.IGNORECASE)
    line = re.sub(r"\s+([,.;:!?])", r"\1", line)
    line = re.sub(r"([\[(])\s+", r"\1", line)
    line = re.sub(r"\s+([\])])", r"\1", line)
    line = re.sub(r"\s+-\s+", " - ", line)
    return line.strip()


def clean_prompt_text(prompt: str) -> str:
    cleaned_lines: list[str] = []
    seen: set[str] = set()
    for raw_line in prompt.splitlines():
        line = clean_prompt_line(raw_line)
        if not line:
            continue
        key = line.casefold()
        if key in seen:
            continue
        seen.add(key)
        cleaned_lines.append(line)
    return "\n".join(cleaned_lines).strip()


def normalize_result(data: dict, frame_paths: list[Path]) -> dict:
    prompt_lines = [clean_prompt_line(str(line).rstrip()) for line in data.get("prompt_lines", []) if str(line).strip()]
    if not prompt_lines:
        fallback_lines = [clean_prompt_line(str(line).rstrip()) for line in data.get("candidate_lines", []) if str(line).strip()]
        prompt_lines = fallback_lines
    prompt = clean_prompt_text("\n".join(prompt_lines).strip())
    cleaned_lines = [line for line in prompt.splitlines() if line.strip()]
    image_frame = frame_paths[len(frame_paths) // 2].name if frame_paths else ""
    return {
        "success": bool(prompt),
        "prompt": prompt,
        "ko": "",
        "camera": detect_camera_movement(prompt),
        "summary_slug": sanitize_slug(cleaned_lines[0] if cleaned_lines else "", fallback="untitled_scene"),
        "image_frame": image_frame,
    }


def finalize_result(raw: dict, frame_paths: list[Path], ocr_results: list[dict]) -> dict:
    result = normalize_result(raw, frame_paths)
    if result.get("success"):
        return result
    fallback_lines = fallback_prompt_lines_from_ocr(ocr_results)
    if fallback_lines:
        result = normalize_result({"prompt_lines": fallback_lines}, frame_paths)
    return result


def find_existing_work_dir(video_id: str, frames_dir: Path) -> Path | None:
    direct = frames_dir / video_id
    if direct.exists():
        return direct
    for path in frames_dir.glob(f"*_{video_id}"):
        if path.is_dir():
            return path
    return None


def find_existing_video(work_dir: Path) -> Path | None:
    videos = sorted(work_dir.glob("video.*"))
    return videos[0] if videos else None


def find_existing_frames(work_dir: Path) -> list[Path]:
    return sorted(work_dir.glob("frame_*.jpg"))


def find_image_frames(work_dir: Path) -> list[Path]:
    candidates: list[Path] = []
    for pattern in (
        "frame_*.jpg",
        "frame_*.jpeg",
        "frame_*.png",
        "frame_*.webp",
        "frame_*.JPG",
        "frame_*.JPEG",
        "frame_*.PNG",
        "frame_*.WEBP",
        "*.jpg",
        "*.jpeg",
        "*.png",
        "*.webp",
        "*.JPG",
        "*.JPEG",
        "*.PNG",
        "*.WEBP",
    ):
        candidates.extend(work_dir.glob(pattern))
    unique = sorted({path.resolve(): path for path in candidates}.values(), key=lambda path: path.name.lower())
    return unique


def build_best_effort_result(
    frame_paths: list[Path],
    ocr_results: list[dict],
    aggregated_ocr: str,
    raw: dict | None = None,
    gemini_error: str = "",
) -> tuple[dict, str]:
    raw = raw or {}
    result = finalize_result(raw, frame_paths, ocr_results)
    if result.get("success"):
        return result, gemini_error

    fallback_lines = fallback_prompt_lines_from_ocr(ocr_results)
    if fallback_lines:
        return normalize_result({"prompt_lines": fallback_lines}, frame_paths), gemini_error or "fallback_prompt_lines_from_ocr 사용"

    aggregated_lines = [line.strip() for line in aggregated_ocr.splitlines() if line.strip()]
    if aggregated_lines:
        return normalize_result({"prompt_lines": aggregated_lines[:20]}, frame_paths), gemini_error or "OCR 집계 텍스트 사용"

    return normalize_result({}, frame_paths), gemini_error


def ensure_unique_dir_name(base_slug: str, video_id: str, frames_dir: Path) -> Path:
    target = frames_dir / f"{base_slug}_{video_id}"
    if not target.exists():
        return target
    suffix = 2
    while True:
        candidate = frames_dir / f"{base_slug}_{video_id}_{suffix}"
        if not candidate.exists():
            return candidate
        suffix += 1


def rename_work_dir(work_dir: Path, slug: str, video_id: str, frames_dir: Path) -> Path:
    desired = ensure_unique_dir_name(slug, video_id, frames_dir)
    if work_dir.resolve() == desired.resolve():
        return work_dir
    work_dir.rename(desired)
    return desired


def create_thumbnail(source_frame: Path, work_dir: Path) -> Path:
    thumbnail_path = work_dir / THUMBNAIL_NAME
    with PILImage.open(source_frame) as image:
        image = image.convert("RGB")
        image.thumbnail((480, 480))
        image.save(thumbnail_path, format="JPEG", quality=90)
    return thumbnail_path


def upsert_csv_row(row: dict, csv_path: Path) -> None:
    if not row.get("URL", "").strip():
        return
    rows: list[dict[str, str]] = []
    if csv_path.exists():
        with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows.extend(csv.DictReader(handle))

    export = export_row(row)
    replaced = False
    row_key = export["링크"]
    for index, existing in enumerate(rows):
        existing_key = existing.get("링크", "")
        if existing_key == row_key:
            rows[index] = export
            replaced = True
            break
    if not replaced:
        rows.append(export)

    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=DEFAULT_HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(csv_path: Path, xlsx_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"
    sheet.append(DEFAULT_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {"A": 24, "B": 72, "C": 40, "D": 42}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    if csv_path.exists():
        with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
        for row_index, row in enumerate(rows, start=2):
            for col_index, header in enumerate(DEFAULT_HEADERS, start=1):
                sheet.cell(row_index, col_index, row.get(header, ""))
                sheet.cell(row_index, col_index).alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(xlsx_path)


def sync_to_google_sheet(row: dict, config: dict) -> str:
    if not row.get("URL", "").strip():
        return "링크 없음: Google Sheets 동기화 건너뜀"
    google_sheet_id = str(config.get("google_sheet_id", "")).strip()
    worksheet_name = str(config.get("worksheet", DEFAULT_WORKSHEET)).strip() or DEFAULT_WORKSHEET
    credentials_path = credentials_path_for(config)

    if not google_sheet_id:
        return "Google Sheet ID 미설정"
    if not credentials_path.exists():
        return f"credentials.json 없음: {credentials_path}"

    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        return "gspread/google-auth 미설치"

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    try:
        creds = Credentials.from_service_account_file(str(credentials_path), scopes=scopes)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(google_sheet_id)
        try:
            ws = sh.worksheet(worksheet_name)
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(title=worksheet_name, rows=1000, cols=20)

        values = ws.get_all_values()
    except Exception as exc:
        return f"Google Sheets 인증/접속 오류: {exc}"
    if not values:
        ws.append_row(DEFAULT_HEADERS, value_input_option="USER_ENTERED")
        values = [DEFAULT_HEADERS]
    elif values[0] != DEFAULT_HEADERS:
        ws.update(values=[DEFAULT_HEADERS], range_name="A1:D1", value_input_option="USER_ENTERED")

    url_value = row["URL"].strip()
    sheet_link_value = (
        f'=HYPERLINK("{url_value}","{url_value}")'
        if url_value.startswith(("http://", "https://"))
        else url_value
    )
    sheet_row = [[
        row["CAMERA"],
        row["PROMPT"],
        row["KO"],
        sheet_link_value,
    ]]

    existing_urls = ws.col_values(4)
    target_row = None
    for index, value in enumerate(existing_urls[1:], start=2):
        if row["URL"] == value:
            target_row = index
            break

    if target_row:
        ws.update(values=sheet_row, range_name=f"A{target_row}:D{target_row}", value_input_option="USER_ENTERED")
    else:
        ws.append_row(sheet_row[0], value_input_option="USER_ENTERED")

    ws.freeze(rows=1)
    sh.batch_update(
        {
            "requests": [
                {
                    "setBasicFilter": {
                        "filter": {
                            "range": {
                                "sheetId": ws.id,
                                "startRowIndex": 0,
                                "startColumnIndex": 0,
                                "endColumnIndex": 4,
                            }
                        }
                    }
                }
            ]
        }
    )
    return "Google Sheets 동기화 완료"


def delete_result_by_url(url: str, config_overrides: dict | None = None) -> dict:
    config = load_config()
    if config_overrides:
        config.update({k: v for k, v in config_overrides.items() if v not in ("", None)})
    config = save_config(config)

    csv_path = csv_path_for(config)
    xlsx_path = xlsx_path_for(config)
    removed = False
    rows: list[dict[str, str]] = []
    if csv_path.exists():
        with csv_path.open('r', encoding='utf-8-sig', newline='') as handle:
            existing_rows = list(csv.DictReader(handle))
        for row in existing_rows:
            if row.get('링크', '').strip() == url.strip():
                removed = True
                continue
            rows.append({header: row.get(header, '') for header in DEFAULT_HEADERS})
        with csv_path.open('w', encoding='utf-8-sig', newline='') as handle:
            writer = csv.DictWriter(handle, fieldnames=DEFAULT_HEADERS)
            writer.writeheader()
            writer.writerows(rows)
        write_xlsx(csv_path, xlsx_path)
    delete_result_metadata(url, config)

    sheet_status = delete_google_sheet_row(url, config)
    return {
        'success': removed or sheet_status.startswith('Google Sheets'),
        'removed': removed,
        'sheet_status': sheet_status,
    }


def clear_all_results(config_overrides: dict | None = None) -> dict:
    config = load_config()
    if config_overrides:
        config.update({k: v for k, v in config_overrides.items() if v not in ("", None)})
    config = save_config(config)

    csv_path = csv_path_for(config)
    xlsx_path = xlsx_path_for(config)
    with csv_path.open('w', encoding='utf-8-sig', newline='') as handle:
        writer = csv.DictWriter(handle, fieldnames=DEFAULT_HEADERS)
        writer.writeheader()
    write_xlsx(csv_path, xlsx_path)
    clear_results_metadata(config)
    sheet_status = clear_google_sheet_results(config)
    return {'success': True, 'sheet_status': sheet_status}


def save_generated_result(
    prompt: str,
    summary: str = "",
    camera_tags: list[str] | None = None,
    metadata: dict | None = None,
    config_overrides: dict | None = None,
) -> dict:
    config = load_config()
    if config_overrides:
        config.update({k: v for k, v in config_overrides.items() if v not in ("", None)})
    config = save_config(config)

    csv_path = csv_path_for(config)
    xlsx_path = xlsx_path_for(config)
    unique_tags: list[str] = []
    for tag in camera_tags or []:
        clean = " ".join(str(tag).split()).strip()
        if clean and clean not in unique_tags:
            unique_tags.append(clean)
    synthetic_url = f"grommy://generated/{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    raw_row = {
        "URL": synthetic_url,
        "PROMPT": (prompt or "").strip(),
        "KO": (summary or "").strip(),
        "CONTENT": (summary or "").strip(),
        "CAMERA": " / ".join(unique_tags),
    }
    upsert_csv_row(raw_row, csv_path)
    write_xlsx(csv_path, xlsx_path)
    generated_meta = {
        "source": "generated",
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "camera_tags": unique_tags,
    }
    if isinstance(metadata, dict):
        generated_meta.update({key: value for key, value in metadata.items() if value not in ("", None, [], {})})
    upsert_result_metadata(synthetic_url, generated_meta, config)
    return {
        "success": True,
        "sheet_status": "생성 초안을 스튜디오 이력에 저장했습니다. 원본 추출 아카이브와 Google Sheet에는 섞지 않습니다.",
        "row": export_row(raw_row),
    }


def delete_google_sheet_row(url: str, config: dict) -> str:
    google_sheet_id = str(config.get('google_sheet_id', '')).strip()
    worksheet_name = str(config.get('worksheet', DEFAULT_WORKSHEET)).strip() or DEFAULT_WORKSHEET
    credentials_path = credentials_path_for(config)
    if not google_sheet_id:
        return 'Google Sheet ID 미설정'
    if not credentials_path.exists():
        return f'credentials.json 없음: {credentials_path}'
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        return 'gspread/google-auth 미설치'

    scopes = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive',
    ]
    try:
        creds = Credentials.from_service_account_file(str(credentials_path), scopes=scopes)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(google_sheet_id)
        ws = sh.worksheet(worksheet_name)
        values = ws.get_all_values()
    except Exception as exc:
        return f'Google Sheets 인증/접속 오류: {exc}'

    for index, row in enumerate(values[1:], start=2):
        if len(row) >= 4 and row[3].strip() == url.strip():
            ws.delete_rows(index)
            return 'Google Sheets 항목 삭제 완료'
    return 'Google Sheets에서 일치 항목 없음'


def clear_google_sheet_results(config: dict) -> str:
    google_sheet_id = str(config.get('google_sheet_id', '')).strip()
    worksheet_name = str(config.get('worksheet', DEFAULT_WORKSHEET)).strip() or DEFAULT_WORKSHEET
    credentials_path = credentials_path_for(config)
    if not google_sheet_id:
        return 'Google Sheet ID 미설정'
    if not credentials_path.exists():
        return f'credentials.json 없음: {credentials_path}'
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        return 'gspread/google-auth 미설치'

    scopes = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive',
    ]
    try:
        creds = Credentials.from_service_account_file(str(credentials_path), scopes=scopes)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(google_sheet_id)
        ws = sh.worksheet(worksheet_name)
        ws.clear()
        ws.update(range_name='A1:D1', values=[DEFAULT_HEADERS], value_input_option='USER_ENTERED')
        ws.freeze(rows=1)
        sh.batch_update(
            {
                'requests': [
                    {
                        'setBasicFilter': {
                            'filter': {
                                'range': {
                                    'sheetId': ws.id,
                                    'startRowIndex': 0,
                                    'startColumnIndex': 0,
                                    'endColumnIndex': 4,
                                }
                            }
                        }
                    }
                ]
            }
        )
        return 'Google Sheets 전체 삭제 완료'
    except Exception as exc:
        return f'Google Sheets 인증/접속 오류: {exc}'


def run_pipeline(
    url: str,
    config_overrides: dict | None = None,
    force_reanalyse: bool = False,
    cancel_event: threading.Event | None = None,
) -> dict:
    config = load_config()
    if config_overrides:
        config.update({k: v for k, v in config_overrides.items() if v not in ("", None)})
    config = save_config(config)

    _, _, outputs_dir = ensure_runtime_dirs(config)
    csv_path = outputs_dir / "results.csv"
    xlsx_path = outputs_dir / "results.xlsx"

    video_id = get_video_id(url)
    result_error = ""
    gemini_usage = estimate_gemini_cost("gemini-2.5-flash", 0, 0)
    image_path = ""

    with tempfile.TemporaryDirectory(prefix=f"prompt_extract_{video_id}_") as temp_dir_str:
        work_dir = Path(temp_dir_str)
        video_path = download_video(url, work_dir, cancel_event=cancel_event)
        if not video_path:
            return {
                "success": False,
                "url": url,
                "error": "사용자 중단" if was_cancelled(cancel_event) else f"다운로드 실패: {LAST_DOWNLOAD_ERROR or '원인 미상'}",
            }

        print("🤖 Gemini 영상 직접 분석 중...")
        raw, error, gemini_usage = call_gemini_on_video(video_path, get_gemini_api_key(), cancel_event=cancel_event)
        if error:
            print(f"⚠️  Gemini video error: {error}")
            result = normalize_result(raw, [])
            result_error = error
        else:
            result = normalize_result(raw, [])

        if not result.get("success"):
            error_message = result_error or "Gemini가 영상에서 프롬프트를 찾지 못했습니다."
            lowered = error_message.lower()
            if '404' in lowered or 'not_found' in lowered or 'no longer available' in lowered:
                error_message = 'Gemini 모델 호출에 실패했습니다. 모델 설정을 최신 상태로 다시 시도해 주세요.'
            return {
                "success": False,
                "url": url,
                "prompt": "",
                "error": error_message,
                "gemini_usage": gemini_usage,
            }

        if not result.get("ko"):
            result["ko"] = build_korean_summary({
                "KO": result.get("prompt", ""),
                "PROMPT": result.get("prompt", ""),
                "URL": url,
            })

        row = {
            "URL": url,
            "IMAGE": image_path,
            "PROMPT": result["prompt"],
            "KO": result["ko"],
            "CAMERA": result["camera"],
        }
        upsert_csv_row(row, csv_path)
        write_xlsx(csv_path, xlsx_path)
        upsert_result_metadata(
            url,
            {
                "source": "extract",
                "saved_at": datetime.now().isoformat(timespec="seconds"),
                "gemini_usage": gemini_usage,
                "camera_tags": [tag.strip() for tag in str(result.get("camera", "")).split("/") if tag.strip()],
                "warning": result_error or "",
            },
            config,
        )
        sheet_status = sync_to_google_sheet(row, config)

        return {
            "success": True,
            "url": url,
            "image_path": image_path,
            "prompt": result["prompt"],
            "ko": result["ko"],
            "camera": result["camera"],
            "work_dir": "",
            "csv_path": str(csv_path),
            "xlsx_path": str(xlsx_path),
            "sheet_status": sheet_status,
            "config_path": str(CONFIG_PATH),
            "error": result_error,
            "gemini_usage": gemini_usage,
        }


def run_frames_only(
    url: str,
    config_overrides: dict | None = None,
    force_reextract: bool = False,
    cancel_event: threading.Event | None = None,
) -> dict:
    config = load_config()
    if config_overrides:
        config.update({k: v for k, v in config_overrides.items() if v not in ("", None)})
    config = save_config(config)

    _, frames_dir, _ = ensure_runtime_dirs(config)
    video_id = get_video_id(url)
    work_dir = find_existing_work_dir(video_id, frames_dir) or (frames_dir / video_id)
    work_dir.mkdir(parents=True, exist_ok=True)

    frame_paths = find_existing_frames(work_dir)
    video_path = find_existing_video(work_dir)

    if frame_paths and not force_reextract:
        print(f"ℹ️  기존 프레임 {len(frame_paths)}장 재사용: {work_dir}")
        return {
            "success": True,
            "url": url,
            "video_path": str(video_path) if video_path else "",
            "frame_count": len(frame_paths),
            "work_dir": str(work_dir),
            "message": "기존 프레임 재사용 완료",
        }

    if not video_path:
        video_path = download_video(url, work_dir, cancel_event=cancel_event)
        if not video_path:
            return {
                "success": False,
                "url": url,
                "error": "사용자 중단" if was_cancelled(cancel_event) else f"다운로드 실패: {LAST_DOWNLOAD_ERROR or '원인 미상'}",
            }

    frame_paths = extract_frames(video_path, work_dir, cancel_event=cancel_event)
    if not frame_paths:
        return {"success": False, "url": url, "error": "사용자 중단" if was_cancelled(cancel_event) else "프레임 추출 실패"}

    return {
        "success": True,
        "url": url,
        "video_path": str(video_path),
        "frame_count": len(frame_paths),
        "work_dir": str(work_dir),
        "message": "영상 및 프레임 추출 완료",
    }


def run_pipeline_from_frame_dir(
    frame_dir: str | Path,
    config_overrides: dict | None = None,
    url: str = "",
    cancel_event: threading.Event | None = None,
) -> dict:
    config = load_config()
    if config_overrides:
        config.update({k: v for k, v in config_overrides.items() if v not in ("", None)})
    config = save_config(config)

    frame_dir = Path(frame_dir).expanduser().resolve()
    _, _, outputs_dir = ensure_runtime_dirs(config)
    csv_path = outputs_dir / "results.csv"
    xlsx_path = outputs_dir / "results.xlsx"

    if not frame_dir.exists() or not frame_dir.is_dir():
        return {"success": False, "url": url, "prompt": "", "error": "프레임 폴더를 찾을 수 없습니다."}

    frame_paths = find_image_frames(frame_dir)
    if not frame_paths:
        print("❌ no frames")
        return {"success": False, "url": url, "prompt": "", "error": "no frames"}

    print(f"ℹ️  기존 프레임 {len(frame_paths)}장 분석: {frame_dir}")
    ocr_results = run_ocr_all_frames(frame_paths, cancel_event=cancel_event)
    if was_cancelled(cancel_event):
        return {"success": False, "url": url, "prompt": "", "error": "사용자 중단"}

    aggregated_ocr = aggregate_ocr_text(ocr_results)
    if not aggregated_ocr.strip():
        print("❌ OCR empty")
        return {"success": False, "url": url, "prompt": "", "error": "OCR empty"}
    print("🤖 Gemini 분석 중...")
    raw, error, gemini_usage = call_gemini(aggregated_ocr, get_gemini_api_key(), cancel_event=cancel_event)
    if error:
        print(f"⚠️  Gemini error: {error}")
    result, result_error = build_best_effort_result(frame_paths, ocr_results, aggregated_ocr, raw, error)
    if not result.get("success"):
        return {
            "success": False,
            "url": url,
            "prompt": "",
            "error": result_error or "Gemini가 프롬프트 줄을 찾지 못했습니다.",
        }

    source_frame = next((path for path in frame_paths if path.name == result["image_frame"]), frame_paths[len(frame_paths) // 2])
    image_path = create_thumbnail(source_frame, frame_dir)

    if not result.get("ko"):
        result["ko"] = build_korean_summary({
            "KO": build_summary_source(aggregated_ocr, result.get("prompt", "")),
            "PROMPT": result.get("prompt", ""),
            "URL": url,
        })

    row = {
        "URL": url,
        "IMAGE": str(image_path),
        "PROMPT": result["prompt"],
        "KO": result["ko"],
        "CAMERA": result["camera"],
    }
    upsert_csv_row(row, csv_path)
    write_xlsx(csv_path, xlsx_path)
    sheet_status = sync_to_google_sheet(row, config) if url else "URL 매칭 없음: Google Sheets 동기화 건너뜀"

    return {
        "success": True,
        "url": url,
        "image_path": str(image_path),
        "prompt": result["prompt"],
        "ko": result["ko"],
        "camera": result["camera"],
        "work_dir": str(frame_dir),
        "csv_path": str(csv_path),
        "xlsx_path": str(xlsx_path),
        "sheet_status": sheet_status,
        "config_path": str(CONFIG_PATH),
        "error": result_error,
        "gemini_usage": gemini_usage,
    }


def print_summary(result: dict) -> None:
    print("\n" + "=" * 60)
    if not result.get("success"):
        print("❌ 분석 실패")
        print(f"URL   : {result.get('url', '')}")
        print(f"error : {result.get('error', '원인 미상')}")
        print("=" * 60)
        return
    print("✅ 분석 완료")
    print(f"URL        : {result.get('url', '')}")
    print(f"IMAGE      : {result.get('image_path', '')}")
    print(f"CSV        : {result.get('csv_path', '')}")
    print(f"XLSX       : {result.get('xlsx_path', '')}")
    print(f"시트동기화 : {result.get('sheet_status', '')}")
    print("=" * 60)


def main() -> None:
    parser = argparse.ArgumentParser(description="프롬프트 추출 자동화")
    parser.add_argument("url")
    parser.add_argument("--reanalyse", action="store_true")
    args = parser.parse_args()
    result = run_pipeline(args.url, force_reanalyse=args.reanalyse)
    print_summary(result)
    if not result.get("success"):
        sys.exit(1)


if __name__ == "__main__":
    main()
